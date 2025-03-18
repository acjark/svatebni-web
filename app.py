from flask import Flask, request, redirect, render_template, send_from_directory
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = "responses.xlsx"

# Serve static files (HTML, CSS, JS)
@app.route("/")
def home():
    # Load the Excel file and get the "sweets" column
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
        sweets_responses = [row[4] for row in sheet.iter_rows(min_row=2, values_only=True) if row[4]]  # Assuming "sweets" is the 5th column
    else:
        sweets_responses = []

    return render_template("my k vam.html", sweets_responses=sweets_responses)

@app.route("/submit_form", methods=["POST"])
def submit_form():
    # Get form data
    name = request.form.get("name")
    email = request.form.get("email")
    question1 = request.form.get("question1")
    question2 = request.form.get("question2")
    question3 = request.form.get("question3")
    sweets = request.form.get("sweets")  # New field for sweets/baked goods

    # Create a new row of data
    data = [name, email, question1, question2, question3, sweets]

    # Check if the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        # Create a new Excel file and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Jméno", "Email", "Otázka 1", "Otázka 2", "Otázka 3", "Sweets"])
    else:
        # Load the existing Excel file
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

    # Append the new data to the sheet
    sheet.append(data)

    # Save the workbook
    workbook.save(EXCEL_FILE)

    # Redirect back to the form with a success message
    return redirect("/?status=success")

if __name__ == "__main__":
    app.run(debug=True)